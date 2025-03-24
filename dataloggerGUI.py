import sys
import os
import json
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QLabel, QLineEdit, QComboBox, QPushButton, QScrollArea,
                             QMessageBox, QGridLayout, QTabWidget, QFileDialog,
                             QFrame, QListView)
from PyQt6.QtCore import Qt, QTimer, QEvent
from PyQt6.QtGui import QPalette, QColor, QCursor


class FocusLineEdit(QLineEdit):
    """Custom QLineEdit with enhanced focus visualization"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setStyleSheet("""
            QLineEdit {
                border: 1px solid #ababab;
                border-radius: 3px;
                padding: 2px;
                background-color: white;
            }
            QLineEdit:focus {
                border: 2px solid #0078d7;
                background-color: #e5f1fb;
            }
        """)


class FocusComboBox(QComboBox):
    """Custom QComboBox with enhanced focus visualization"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setView(QListView())

        # Set stylesheet with proper colors for all states
        self.setStyleSheet("""
            QComboBox {
                border: 1px solid #ababab;
                border-radius: 3px;
                padding: 1px 18px 1px 3px;
                min-width: 6em;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #0078d7;
                background-color: #e5f1fb;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 15px;
                border-left-width: 1px;
                border-left-color: darkgray;
                border-left-style: solid;
                border-top-right-radius: 3px;
                border-bottom-right-radius: 3px;
            }

            /* Important: Style both the QListView and the items */
            QComboBox QAbstractItemView {
                background: white;
                border: 1px solid #ababab;
                selection-background-color: #0078d7;
                selection-color: white;
            }

            QComboBox QAbstractItemView::item {
                background-color: white;
                color: black;
                padding: 4px;
            }

            QComboBox QAbstractItemView::item:hover {
                background-color: #e5f1fb;
                color: black;
            }

            QComboBox QAbstractItemView::item:selected {
                background-color: #0078d7;
                color: white;
            }
        """)


class DataLogGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config_dir = os.path.join(os.path.expanduser('~'), 'Library', 'Application Support', 'DataLogApp')
        os.makedirs(self.config_dir, exist_ok=True)
        self.config_file = os.path.join(self.config_dir, 'config.json')
        self.file_location = None
        self.workbook_path = None  # Will be set when user chooses location
        self.setWindowTitle("Krienen Data Logger")

        # Initialize these values early as they're lightweight
        self.name_to_code = {
            "Croissant": "CJ23.56.002",
            "Nutmeg": "CJ23.56.003",
            "Jellybean": "CJ24.56.001",
            "Rambo": "CJ24.56.004",
            "Morel": "CJ24.56.015"
        }

        self.tile_location_map = {
            "BRAINSTEM": "BS",
            "BS": "BS",
            "CORTEX": "CX",
            "CX": "CX",
            "CEREBELLUM": "CB",
            "CB": "CB"
        }

        # Show UI right away, defer other initialization
        self.init_ui()

        # Initialize heavy components after UI is visible
        QTimer.singleShot(0, self.delayed_init)

    def delayed_init(self):
        """Initialize components that aren't needed for initial UI display"""
        # Import heavy modules only when needed
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter

        if getattr(sys, 'frozen', False):
            self.script_dir = os.path.dirname(sys.executable)
        else:
            self.script_dir = os.path.dirname(os.path.abspath(__file__))

        self.COUNTER_FILE = os.path.join(self.script_dir, 'sample_name_counter.json')
        self.black_fill = PatternFill(start_color='000000', fill_type='solid')
        self.default_font = Font(name="Arial", size=10)
        self.bold_font = Font(name="Arial", size=10, bold=True)

        # Load counter data in the background
        self.load_counter_data()

        # Setup enter key navigation after UI is completely initialized
        self.setup_enter_key_navigation()

    def get_save_location(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f:
                config = json.load(f)
                return config.get('file_location')

        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.expanduser('~'),  # Start in the user's home directory
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if file_name:
            try:
                with open(self.config_file, 'w') as f:
                    json.dump({'file_location': file_name}, f)
            except IOError as e:
                QMessageBox.critical(self, "Error", f"Failed to write config file: {str(e)}")

        return file_name

    def save_data(self):
        # Import pandas only when needed
        import pandas as pd

        if not self.file_location:
            self.file_location = self.get_save_location()

        if not self.file_location:
            QMessageBox.critical(self, "Error", "No save location specified!")
            return

        file_location = self.file_location
        if not file_location.endswith('.xlsx'):
            file_location += '.xlsx'

        try:
            # Create DataFrame from your data
            data = {
                'krienen_lab_identifier': [self.krienen_lab_identifier],
                'seq_portal': [self.seq_portal],
                # ... add all your other fields here ...
                'Current Date and Time (UTC)': ["2025-03-24 15:02:44"],
                'Current User Login': ["lakmecaceres"]
            }
            df = pd.DataFrame(data)
            df.to_excel(file_location, index=False)
            QMessageBox.information(self, "Success", "Data saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file: {str(e)}")

    def get_current_time(self):
        # Return fixed time as provided
        return "2025-03-24 15:02:44"

    def get_current_user(self):
        # Return fixed user as provided
        return "lakmecaceres"

    def init_ui(self):
        self.setGeometry(100, 100, 800, 600)

        # Create main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)

        # Add a frame with a title to make it clear what the app does
        title_frame = QFrame()
        title_frame.setFrameShape(QFrame.Shape.StyledPanel)
        title_frame.setStyleSheet("background-color: #f0f0f0; padding: 5px;")
        title_layout = QVBoxLayout(title_frame)
        title_label = QLabel("Krienen Lab Data Logger")
        title_label.setStyleSheet("font-size: 16pt; font-weight: bold;")
        title_layout.addWidget(title_label)
        main_layout.addWidget(title_frame)

        # Create tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)

        # Create tabs
        tissue_tab = QWidget()
        facs_tab = QWidget()
        library_tab = QWidget()
        indices_tab = QWidget()

        # Setup layouts for each tab
        self.setup_basic_tab(tissue_tab)
        self.setup_facs_tab(facs_tab)
        self.setup_library_tab(library_tab)
        self.setup_indices_tab(indices_tab)

        # Add tabs to widget with new names
        self.tab_widget.addTab(tissue_tab, "Tissue")
        self.tab_widget.addTab(facs_tab, "FACS")
        self.tab_widget.addTab(library_tab, "cDNA")
        self.tab_widget.addTab(indices_tab, "Libraries")

        # Add submit button
        self.submit_btn = QPushButton('Submit')
        self.submit_btn.clicked.connect(self.on_submit)
        self.submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d7;
                color: white;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #0063b1;
            }
            QPushButton:pressed {
                background-color: #004e8c;
            }
        """)
        main_layout.addWidget(self.submit_btn, alignment=Qt.AlignmentFlag.AlignCenter)

    def setup_enter_key_navigation(self):
        """Set up navigation to next field when Enter key is pressed"""
        # Create a list of all input widgets
        self.input_widgets = []

        # Add widgets from all tabs
        for tab_index in range(self.tab_widget.count()):
            tab = self.tab_widget.widget(tab_index)
            for child in tab.findChildren(QWidget):
                if isinstance(child, QLineEdit):
                    self.input_widgets.append(child)
                    child.returnPressed.connect(self.on_return_pressed)
                elif isinstance(child, QComboBox):
                    self.input_widgets.append(child)
                    child.installEventFilter(self)

    def on_return_pressed(self):
        """Handle Return key press for any input widget"""
        sender = self.sender()
        if sender in self.input_widgets:
            self.move_to_next_widget(sender)

    def eventFilter(self, obj, event):
        """Event filter to handle Enter key in QComboBox"""
        if (event.type() == QEvent.Type.KeyPress and
                isinstance(obj, QComboBox) and
                event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter)):
            self.move_to_next_widget(obj)
            return True

        return super().eventFilter(obj, event)

    def move_to_next_widget(self, current_widget):
        """Move focus to the next widget in the input widgets list"""
        try:
            current_index = self.input_widgets.index(current_widget)
            next_index = (current_index + 1) % len(self.input_widgets)
            next_widget = self.input_widgets[next_index]

            # If we're moving to a widget on a different tab, switch to that tab
            for tab_index in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(tab_index)
                if self.widget_is_in_tab(next_widget, tab):
                    if tab_index != self.tab_widget.currentIndex():
                        self.tab_widget.setCurrentIndex(tab_index)
                    break

            next_widget.setFocus()
        except (ValueError, IndexError) as e:
            print(f"Navigation error: {e}")
            # If widget is not in list or there's an error, do nothing
            pass

    def widget_is_in_tab(self, widget, tab):
        """Check if a widget is contained within a tab"""
        if widget is tab:
            return True

        # Check if widget is directly in tab's children
        if widget in tab.findChildren(widget.__class__):
            return True

        # Check parent chain
        parent = widget.parent()
        while parent:
            if parent is tab:
                return True
            parent = parent.parent()

        return False

    def setup_basic_tab(self, tab):
        layout = QGridLayout()

        # Project selection (at top)
        self.project_input = FocusComboBox()
        self.project_input.addItems(["HMBA_CjAtlas_Subcortex", "Other"])
        self.project_input.currentTextChanged.connect(self.on_project_change)
        layout.addWidget(QLabel("Project:"), 0, 0)
        layout.addWidget(self.project_input, 0, 1)
        self.project_name_input = FocusLineEdit()
        self.project_name_input.setVisible(False)
        layout.addWidget(self.project_name_input, 0, 2)

        # Date input with validation
        self.date_input = FocusLineEdit()
        self.date_input.setPlaceholderText("YYMMDD or MM/DD/YY")
        layout.addWidget(QLabel("Experiment Date:"), 1, 0)
        layout.addWidget(self.date_input, 1, 1)

        # Marmoset name with dropdown
        self.marmoset_input = FocusComboBox()
        self.marmoset_input.addItems(self.name_to_code.keys())
        layout.addWidget(QLabel("Marmoset Name:"), 2, 0)
        layout.addWidget(self.marmoset_input, 2, 1)

        # Hemisphere selection
        self.hemisphere_input = FocusComboBox()
        self.hemisphere_input.addItems(["Left (LH)", "Right (RH)", "Both"])
        layout.addWidget(QLabel("Hemisphere:"), 3, 0)
        layout.addWidget(self.hemisphere_input, 3, 1)

        # Tile location as a simple dropdown (NOT editable)
        self.tile_location_input = FocusComboBox()
        self.tile_location_input.addItems(["BS", "CX", "CB"])
        # No longer editable - making it a pure dropdown
        layout.addWidget(QLabel("Tile Location:"), 4, 0)
        layout.addWidget(self.tile_location_input, 4, 1)

        # Slab and tile numbers
        self.slab_input = FocusLineEdit()
        self.slab_input.setPlaceholderText("Enter numeric value")
        layout.addWidget(QLabel("Slab Number:"), 5, 0)
        layout.addWidget(self.slab_input, 5, 1)

        self.tile_input = FocusLineEdit()
        self.tile_input.setPlaceholderText("Enter numeric value")
        layout.addWidget(QLabel("Tile Number:"), 6, 0)
        layout.addWidget(self.tile_input, 6, 1)

        # Set layout for this tab
        tab.setLayout(layout)

    def setup_facs_tab(self, tab):
        layout = QGridLayout()

        # Sorter Initials
        self.sorter_initials_input = FocusLineEdit()
        self.sorter_initials_input.setPlaceholderText("Enter sorter's initials")
        layout.addWidget(QLabel("Sorter Initials:"), 0, 0)
        layout.addWidget(self.sorter_initials_input, 0, 1)

        # Sort method
        self.sort_method_input = FocusComboBox()
        self.sort_method_input.addItems(["pooled", "unsorted", "DAPI"])
        self.sort_method_input.currentTextChanged.connect(self.on_sort_method_change)
        layout.addWidget(QLabel("Sort Method:"), 1, 0)
        layout.addWidget(self.sort_method_input, 1, 1)

        # FACS population (moved before number of reactions)
        self.facs_population_input = FocusLineEdit()
        self.facs_population_input.setPlaceholderText("Format: XX/XX/XX (e.g., 70/20/10)")
        layout.addWidget(QLabel("FACS Population:"), 2, 0)
        layout.addWidget(self.facs_population_input, 2, 1)

        # Number of Reactions
        self.rxn_number_input = FocusLineEdit()
        self.rxn_number_input.setPlaceholderText("Enter number of reactions")
        layout.addWidget(QLabel("Number of Reactions:"), 3, 0)
        layout.addWidget(self.rxn_number_input, 3, 1)

        # Expected Recovery
        self.expected_recovery_input = FocusLineEdit()
        layout.addWidget(QLabel("Expected Recovery:"), 4, 0)
        layout.addWidget(self.expected_recovery_input, 4, 1)

        # Nuclei Concentration
        self.nuclei_concentration_input = FocusLineEdit()
        layout.addWidget(QLabel("Nuclei Concentration:"), 5, 0)
        layout.addWidget(self.nuclei_concentration_input, 5, 1)

        # Nuclei Volume
        self.nuclei_volume_input = FocusLineEdit()
        layout.addWidget(QLabel("Nuclei Volume (µL):"), 6, 0)
        layout.addWidget(self.nuclei_volume_input, 6, 1)

        tab.setLayout(layout)

    def setup_library_tab(self, tab):
        layout = QGridLayout()

        # Library dates
        self.cdna_amp_date_input = FocusLineEdit()
        self.cdna_amp_date_input.setPlaceholderText("YYMMDD or MM/DD/YY")
        layout.addWidget(QLabel("cDNA Amplification Date:"), 0, 0)
        layout.addWidget(self.cdna_amp_date_input, 0, 1)

        self.atac_prep_date_input = FocusLineEdit()
        self.atac_prep_date_input.setPlaceholderText("YYMMDD or MM/DD/YY")
        layout.addWidget(QLabel("ATAC Library Prep Date:"), 1, 0)
        layout.addWidget(self.atac_prep_date_input, 1, 1)

        self.rna_prep_date_input = FocusLineEdit()
        self.rna_prep_date_input.setPlaceholderText("YYMMDD or MM/DD/YY")
        layout.addWidget(QLabel("cDNA Library Prep Date:"), 2, 0)
        layout.addWidget(self.rna_prep_date_input, 2, 1)

        # PCR cycles
        self.cdna_pcr_cycles_input = FocusLineEdit()
        self.cdna_pcr_cycles_input.setPlaceholderText("Comma-separated values")
        layout.addWidget(QLabel("cDNA PCR Cycles:"), 3, 0)
        layout.addWidget(self.cdna_pcr_cycles_input, 3, 1)

        # cDNA metrics
        self.cdna_concentration_input = FocusLineEdit()
        self.cdna_concentration_input.setPlaceholderText("Comma-separated values (ng/µL)")
        layout.addWidget(QLabel("cDNA Concentration:"), 4, 0)
        layout.addWidget(self.cdna_concentration_input, 4, 1)

        self.percent_cdna_400bp_input = FocusLineEdit()
        self.percent_cdna_400bp_input.setPlaceholderText("Comma-separated values")
        layout.addWidget(QLabel("Percent cDNA > 400bp:"), 5, 0)
        layout.addWidget(self.percent_cdna_400bp_input, 5, 1)

        tab.setLayout(layout)

    def on_sort_method_change(self, value):
        # Update FACS population field based on sort method
        if value.lower() == "pooled":
            self.facs_population_input.setEnabled(True)
            self.facs_population_input.setPlaceholderText("Format: XX/XX/XX (e.g., 70/20/10)")
        elif value.lower() == "unsorted":
            self.facs_population_input.setEnabled(False)
            self.facs_population_input.setText("no_FACS")
        else:  # DAPI
            self.facs_population_input.setEnabled(False)
            self.facs_population_input.setText("DAPI")

    def on_project_change(self, value):
        self.project_name_input.setVisible(value == "Other")

    def load_counter_data(self):
        if os.path.exists(self.COUNTER_FILE):
            with open(self.COUNTER_FILE, 'r') as f:
                try:
                    self.counter_data = json.load(f)
                except json.JSONDecodeError:
                    self.counter_data = {}
        else:
            self.counter_data = {}

        self.counter_data.setdefault("next_counter", 90)
        self.counter_data.setdefault("date_info", {})
        self.counter_data.setdefault("amp_counter", {})

    def setup_indices_tab(self, tab):
        layout = QGridLayout()

        # ATAC fields grouped together
        self.library_cycles_atac_input = FocusLineEdit()
        self.library_cycles_atac_input.setPlaceholderText("Comma-separated values")
        layout.addWidget(QLabel("ATAC Library Cycles:"), 0, 0)
        layout.addWidget(self.library_cycles_atac_input, 0, 1)

        self.atac_indices_input = FocusLineEdit()
        self.atac_indices_input.setPlaceholderText("Comma-separated values (e.g., D4,E5,F6)")
        layout.addWidget(QLabel("ATAC Indices:"), 1, 0)
        layout.addWidget(self.atac_indices_input, 1, 1)

        self.atac_lib_concentration_input = FocusLineEdit()
        self.atac_lib_concentration_input.setPlaceholderText("Comma-separated values (ng/µL)")
        layout.addWidget(QLabel("ATAC Library Concentration:"), 2, 0)
        layout.addWidget(self.atac_lib_concentration_input, 2, 1)

        self.atac_sizes_input = FocusLineEdit()
        self.atac_sizes_input.setPlaceholderText("Comma-separated values")
        layout.addWidget(QLabel("ATAC Library Sizes (bp):"), 3, 0)
        layout.addWidget(self.atac_sizes_input, 3, 1)

        # cDNA fields grouped together
        self.library_cycles_rna_input = FocusLineEdit()
        self.library_cycles_rna_input.setPlaceholderText("Comma-separated values")
        layout.addWidget(QLabel("cDNA Library Cycles:"), 4, 0)
        layout.addWidget(self.library_cycles_rna_input, 4, 1)

        self.rna_indices_input = FocusLineEdit()
        self.rna_indices_input.setPlaceholderText("Comma-separated values (e.g., A1,B2,C3)")
        layout.addWidget(QLabel("cDNA Indices:"), 5, 0)
        layout.addWidget(self.rna_indices_input, 5, 1)

        self.rna_lib_concentration_input = FocusLineEdit()
        self.rna_lib_concentration_input.setPlaceholderText("Comma-separated values (ng/µL)")
        layout.addWidget(QLabel("cDNA Library Concentration:"), 6, 0)
        layout.addWidget(self.rna_lib_concentration_input, 6, 1)

        self.rna_sizes_input = FocusLineEdit()
        self.rna_sizes_input.setPlaceholderText("Comma-separated values")
        layout.addWidget(QLabel("cDNA Library Sizes (bp):"), 7, 0)
        layout.addWidget(self.rna_sizes_input, 7, 1)

        tab.setLayout(layout)

    def convert_index(self, index):
        index = index.strip().upper()
        if len(index) == 3:
            if index[0].isdigit() and index[1].isdigit() and index[2].isalpha():
                return f"{index[2]}{index[0]}{index[1]}"
            elif index[0].isalpha() and index[1].isdigit() and index[2].isdigit():
                return index
        elif len(index) == 2:
            if index[0].isdigit() and index[1].isalpha():
                return f"{index[1]}0{index[0]}"
            elif index[0].isalpha() and index[1].isdigit():
                return f"{index[0]}0{index[1]}"
        return None

    def pad_index(self, index):
        if len(index) == 2 and index[0].isalpha() and index[1].isdigit():
            return f"{index[0]}0{index[1]}"
        return index

    def convert_date(self, exp_date):
        # Import dateutil only when needed
        import dateutil.parser
        from datetime import datetime

        clean_date = "".join(c for c in exp_date if c.isdigit())
        if len(clean_date) == 6:
            try:
                datetime.strptime(clean_date, '%y%m%d')
                return clean_date
            except ValueError:
                pass
        try:
            parsed_date = dateutil.parser.parse(exp_date)
            return parsed_date.strftime('%y%m%d')
        except ValueError:
            return None

    def validate_inputs(self):
        # Basic validation
        current_date = self.convert_date(self.date_input.text())
        if not current_date:
            QMessageBox.warning(self, "Validation Error", "Please enter a valid date.")
            return False

        try:
            rxn_number = int(self.rxn_number_input.text())
            if rxn_number <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "Validation Error", "Please enter a valid number of reactions.")
            return False

        # Validate numerical inputs
        try:
            slab_int = int(self.slab_input.text())
            tile_int = int(self.tile_input.text())
        except ValueError:
            QMessageBox.warning(self, "Validation Error", "Slab and tile numbers must be numeric values.")
            return False

        # Validate FACS population for pooled samples
        if self.sort_method_input.currentText().lower() == "pooled":
            proportions = self.facs_population_input.text().strip()
            if "/" not in proportions:
                QMessageBox.warning(self, "Validation Error",
                                    "Please enter FACS population proportions in format XX/XX/XX")
                return False
            try:
                proportions_list = [int(p) for p in proportions.split("/")]
                if len(proportions_list) != 3 or sum(proportions_list) != 100:
                    raise ValueError
            except ValueError:
                QMessageBox.warning(self, "Validation Error", "FACS proportions must be three numbers that sum to 100")
                return False

        for field, field_name in [
            (self.percent_cdna_400bp_input, "Percent cDNA > 400bp"),
            (self.cdna_concentration_input, "cDNA concentration"),
            (self.rna_lib_concentration_input, "RNA library concentration"),
            (self.atac_lib_concentration_input, "ATAC library concentration")
        ]:
            values = field.text().strip().split(',')
            try:
                values = [float(x.strip()) for x in values]
                if len(values) != rxn_number:
                    QMessageBox.warning(self, "Validation Error",
                                        f"{field_name} must have {rxn_number} comma-separated values")
                    return False
            except ValueError:
                QMessageBox.warning(self, "Validation Error",
                                    f"{field_name} values must be numbers")
                return False

        # Validate comma-separated inputs match reaction number
        fields_to_validate = [
            (self.cdna_pcr_cycles_input, "cDNA PCR cycles"),
            (self.rna_indices_input, "RNA indices"),
            (self.atac_indices_input, "ATAC indices"),
            (self.rna_sizes_input, "RNA library sizes"),
            (self.atac_sizes_input, "ATAC library sizes"),
            (self.library_cycles_rna_input, "RNA library cycles"),
            (self.library_cycles_atac_input, "ATAC library cycles")
        ]

        for field, field_name in fields_to_validate:
            values = field.text().strip().split(',')
            if len(values) != rxn_number:
                QMessageBox.warning(self, "Validation Error",
                                    f"{field_name} must have {rxn_number} comma-separated values")
                return False

        return True

    def initialize_excel(self):
        # Import openpyxl only when needed
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, NamedStyle

        wb = Workbook()
        ws = wb.active
        ws.title = "HMBA"

        # Create default style (Arial 10)
        default_style = NamedStyle(name="default_style")
        default_style.font = Font(name="Arial", size=10)
        wb.add_named_style(default_style)

        # Create bold style (Arial 10 bold)
        bold_style = NamedStyle(name="bold_style")
        bold_style.font = Font(name="Arial", size=10, bold=True)
        wb.add_named_style(bold_style)

        # Apply default style to entire worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.style = "default_style"

        headers = ['krienen_lab_identifier', 'seq_portal', 'elab_link', 'experiment_start_date',
                   'mit_name', 'donor_name', 'tissue_name', 'tissue_name_old',
                   'dissociated_cell_sample_name', 'facs_population_plan', 'cell_prep_type',
                   'study', 'enriched_cell_sample_container_name', 'expc_cell_capture',
                   'port_well', 'enriched_cell_sample_name', 'enriched_cell_sample_quantity_count',
                   'barcoded_cell_sample_name', 'library_method', 'cDNA_amplification_method',
                   'cDNA_amplification_date', 'amplified_cdna_name', 'cDNA_pcr_cycles',
                   'rna_amplification_pass_fail', 'percent_cdna_longer_than_400bp',
                   'cdna_amplified_quantity_ng', 'cDNA_library_input_ng', 'library_creation_date',
                   'library_prep_set', 'library_name', 'tapestation_avg_size_bp',
                   'library_num_cycles', 'lib_quantification_ng', 'library_prep_pass_fail',
                   'r1_index', 'r2_index', 'ATAC_index']

        ws.append(headers)

        # Apply bold style to headers
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num).style = "bold_style"

        return wb

    def process_form_data(self):
        # Import heavy modules only when needed
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import pyperclip

        # Load or create workbook
        if self.workbook_path and os.path.exists(self.workbook_path):
            workbook = load_workbook(self.workbook_path)
        else:
            workbook = self.initialize_excel()

        worksheet = workbook.active
        current_row = worksheet.max_row
        if current_row == 1 and not any(cell.value for cell in worksheet[1]):
            current_row = 1
        else:
            current_row += 1

        # Get form values
        current_date = self.convert_date(self.date_input.text())
        mit_name_input = self.marmoset_input.currentText()
        mit_name = "cj" + mit_name_input
        donor_name = self.name_to_code[mit_name_input]  # Define donor_name here

        # Process slab and hemisphere
        slab = self.slab_input.text().strip()
        hemisphere = self.hemisphere_input.currentText().split()[0].upper()
        if hemisphere == "RIGHT":
            slab = str(int(slab) + 40).zfill(2)
        elif hemisphere == "BOTH":
            slab = str(int(slab) + 90).zfill(2)
        else:
            slab = slab.zfill(2)

        tile = str(int(self.tile_input.text())).zfill(2)

        # Process tile location
        tile_location_abbr = self.tile_location_input.currentText()

        # Sort method and FACS population
        sort_method = self.sort_method_input.currentText()
        sort_method = sort_method.upper() if sort_method.lower() == "dapi" else sort_method

        if sort_method.lower() == "pooled":
            facs_population = self.facs_population_input.text()
        elif sort_method.lower() == "unsorted":
            facs_population = "no_FACS"
        else:
            facs_population = "DAPI"

        # Get reaction number and update counters
        rxn_number = int(self.rxn_number_input.text())

        # Update date_info
        if current_date not in self.counter_data["date_info"]:
            self.counter_data["date_info"][current_date] = {
                "total_reactions": 0,
                "batches": []
            }

        date_info = self.counter_data["date_info"]
        date_entry = date_info[current_date]
        existing_total = date_entry["total_reactions"]

        # Calculate batch information
        total_reactions_after = existing_total + rxn_number
        batches_before = (existing_total + 7) // 8
        batches_after = (total_reactions_after + 7) // 8
        new_batches_needed = batches_after - batches_before

        new_p_numbers = [self.counter_data["next_counter"] + i for i in range(new_batches_needed)]
        self.counter_data["next_counter"] += new_batches_needed

        all_batches = date_entry["batches"].copy()
        all_batches.extend({"p_number": p, "count": 0} for p in new_p_numbers)

        # Calculate port wells
        port_wells = []
        for x in range(rxn_number):
            global_idx = existing_total + x + 1
            batch_idx = (global_idx - 1) // 8
            p_number = all_batches[batch_idx]["p_number"]
            port_well = (global_idx - 1) % 8 + 1
            port_wells.append((p_number, port_well))

        # Update counters
        date_entry["total_reactions"] = total_reactions_after
        date_entry["batches"] = all_batches

        # Process indices
        atac_indices = [self.convert_index(index) for index in self.atac_indices_input.text().split(",")]
        atac_indices = [self.pad_index(index) for index in atac_indices]

        rna_indices = [self.convert_index(index) for index in self.rna_indices_input.text().split(",")]
        rna_indices = [self.pad_index(index) for index in rna_indices]

        # Initialize common values
        seq_portal = "no"
        elab_link = pyperclip.paste()
        tissue_name = f"{donor_name}.{tile_location_abbr}.{slab}.{tile}"
        dissociated_cell_sample_name = f'{current_date}_{tissue_name}.Multiome'
        cell_prep_type = "nuclei"

        sorting_status = "PS" if sort_method.lower() in ["pooled", "dapi"] else "PN"
        sorter_initials = self.sorter_initials_input.text().strip().upper()
        enriched_cell_sample_container_name = f"MPXM_{current_date}_{sorting_status}_{sorter_initials}"

        # Get study name
        study = "HMBA_CjAtlas_Subcortex" if self.project_input.currentText() == "HMBA_CjAtlas_Subcortex" else self.project_name_input.text()

        # Process the data for each reaction and modality
        dup_index_counter = {}
        headers = [cell.value for cell in worksheet[1]]

        for x in range(rxn_number):
            p_number, port_well = port_wells[x]
            barcoded_cell_sample_name = f'P{str(p_number).zfill(4)}_{port_well}'

            for modality in ["RNA", "ATAC"]:
                self.write_modality_data(
                    worksheet, current_row, modality, x,
                    current_date, mit_name, slab, tile, sort_method,
                    port_well, barcoded_cell_sample_name,
                    sorting_status, sorter_initials,
                    tissue_name, dissociated_cell_sample_name,
                    enriched_cell_sample_container_name,
                    study, seq_portal, elab_link,
                    facs_population, cell_prep_type,
                    rna_indices, atac_indices,
                    headers, dup_index_counter,
                    donor_name  # Add donor_name here
                )
                current_row += 1

        # Save workbook and counter data
        workbook.save(self.workbook_path)
        with open(self.COUNTER_FILE, 'w') as f:
            json.dump(self.counter_data, f, indent=4)

    def write_modality_data(self, worksheet, current_row, modality, x, *args):
        (current_date, mit_name, slab, tile, sort_method,
         port_well, barcoded_cell_sample_name,
         sorting_status, sorter_initials,
         tissue_name, dissociated_cell_sample_name,
         enriched_cell_sample_container_name,
         study, seq_portal, elab_link,
         facs_population, cell_prep_type,
         rna_indices, atac_indices,
         headers, dup_index_counter, donor_name) = args

        krienen_lab_identifier = f'{current_date}_HMBA_{mit_name}_Slab{int(slab)}_Tile{int(tile)}_{sort_method}_{modality}{x + 1}'
        enriched_cell_sample_name = f'MPXM_{current_date}_{sorting_status}_{sorter_initials}_{port_well}'

        library_prep_date = (self.convert_date(self.rna_prep_date_input.text()) if modality == "RNA"
                             else self.convert_date(self.atac_prep_date_input.text()))

        if modality == "RNA":
            library_method = "10xMultiome-RSeq"
            library_type = "LPLCXR"
            library_index = rna_indices[x]

            # Calculate RNA-specific metrics
            cdna_concentration = float(self.cdna_concentration_input.text().split(',')[x])
            cdna_amplified_quantity = cdna_concentration * 40  # 40µL volume for cDNA
            cdna_library_input = cdna_amplified_quantity * 0.25  # 25% of amplified quantity
            percent_cdna_400bp = float(self.percent_cdna_400bp_input.text().split(',')[x])
            rna_concentration = float(self.rna_lib_concentration_input.text().split(',')[x])
            lib_quant = rna_concentration * 35  # Fixed 35µL volume for RNA library

            cdna_pcr_cycles = int(self.cdna_pcr_cycles_input.text().split(',')[x])
            rna_size = int(self.rna_sizes_input.text().split(',')[x])
            library_cycles = int(self.library_cycles_rna_input.text().split(',')[x])
        else:  # ATAC
            library_method = "10xMultiome-ASeq"
            library_type = "LPLCXA"
            library_index = atac_indices[x]

            # Calculate ATAC-specific metrics
            atac_concentration = float(self.atac_lib_concentration_input.text().split(',')[x])
            lib_quant = atac_concentration * 20  # Fixed 20µL volume for ATAC library

            atac_size = int(self.atac_sizes_input.text().split(',')[x])
            library_cycles = int(self.library_cycles_atac_input.text().split(',')[x])

        # Update library prep set counter
        key = (library_type, library_prep_date, library_index)
        dup_index_counter[key] = dup_index_counter.get(key, 0) + 1
        library_prep_set = f"{library_type}_{library_prep_date}_{dup_index_counter[key]}"
        library_name = f"{library_prep_set}_{library_index}"

        # Calculate common metrics
        expected_cell_capture = int(self.expected_recovery_input.text())
        concentration = float(self.nuclei_concentration_input.text().replace(",", ""))
        volume = float(self.nuclei_volume_input.text())
        enriched_cell_sample_quantity_count = round(concentration * volume)

        # Prepare row data
        row_data = [
            krienen_lab_identifier,  # Column 1
            seq_portal,
            elab_link,
            current_date,
            mit_name,
            donor_name,
            tissue_name,
            None,  # tissue_name_old (will be filled black)
            dissociated_cell_sample_name,
            facs_population,
            cell_prep_type,
            study,
            enriched_cell_sample_container_name,
            expected_cell_capture,
            port_well,
            enriched_cell_sample_name,
            enriched_cell_sample_quantity_count,
            barcoded_cell_sample_name,
            library_method,
            "10xMultiome-RSeq" if modality == "RNA" else None,
            self.convert_date(self.cdna_amp_date_input.text()) if modality == "RNA" else None,
            None,  # amplified_cdna_name (filled conditionally)
            cdna_pcr_cycles if modality == "RNA" else None,
            "Pass" if modality == "RNA" else None,
            percent_cdna_400bp if modality == "RNA" else None,
            cdna_amplified_quantity if modality == "RNA" else None,
            cdna_library_input if modality == "RNA" else None,
            library_prep_date,
            library_prep_set,
            library_name,
            rna_size if modality == "RNA" else atac_size,
            library_cycles,
            lib_quant,
            "Pass",
            f"SI-TT-{rna_indices[x]}_i7" if modality == "RNA" else None,
            f"SI-TT-{rna_indices[x]}_b(i5)" if modality == "RNA" else None,
            f"SI-NA-{atac_indices[x]}" if modality == "ATAC" else None
        ]

        # Handle amplified_cdna_name for RNA
        if modality == "RNA":
            if current_date not in self.counter_data["amp_counter"]:
                self.counter_data["amp_counter"][current_date] = 0
            reaction_count = self.counter_data["amp_counter"][current_date]
            letter = chr(65 + (reaction_count % 8))  # A-H (65 is ASCII for 'A')
            batch_num_for_amp = (reaction_count // 8) + 1  # Increment batch number every 8 reactions
            cdna_amp_date = self.convert_date(self.cdna_amp_date_input.text())
            row_data[21] = f"APLCXR_{cdna_amp_date}_{batch_num_for_amp}_{letter}"
            self.counter_data["amp_counter"][current_date] += 1

        # Write to Excel
        for col_num, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=current_row, column=col_num, value=value)
            # Apply default Arial 10 font to all cells
            cell.font = Font(name="Arial", size=10)

            # Apply black fill for ATAC empty cells
            if modality == "ATAC" and value is None:
                cell.fill = self.black_fill

        # Apply black fill to tissue_name_old
        tissue_old_col = headers.index('tissue_name_old') + 1
        worksheet.cell(row=current_row, column=tissue_old_col).fill = self.black_fill

    def on_submit(self):
        try:
            # Change cursor to wait cursor
            QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))

            if not self.validate_inputs():
                QApplication.restoreOverrideCursor()
                return

            # Get file location if not already set
            if not self.file_location:
                QApplication.restoreOverrideCursor()  # Restore cursor before dialog
                self.file_location = self.get_save_location()
                if not self.file_location:
                    QMessageBox.critical(self, "Error", "No save location specified!")
                    return
                QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))  # Set wait cursor again

            # Use file_location instead of workbook_path
            self.workbook_path = self.file_location

            # Process the form data and update Excel
            self.process_form_data()

            # These imports are now inside methods that use them
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            # Adjust column widths
            workbook = load_workbook(self.workbook_path)
            worksheet = workbook.active

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(self.workbook_path)

            # Restore cursor before showing message
            QApplication.restoreOverrideCursor()

            QMessageBox.information(
                self,
                "Success",
                f"Data successfully appended to {self.workbook_path}"
            )

            # Clear form fields after successful submission
            self.clear_form_fields()

        except Exception as e:
            QApplication.restoreOverrideCursor()  # Make sure cursor is restored on error
            QMessageBox.critical(
                self,
                "Error",
                f"An error occurred while processing the data:\n{str(e)}"
            )

    def clear_form_fields(self):
        """Clear all form fields after successful submission"""
        # Clear basic info
        self.date_input.clear()
        self.marmoset_input.setCurrentIndex(0)
        self.slab_input.clear()
        self.tile_input.clear()
        self.hemisphere_input.setCurrentIndex(0)

        # Clear sample info
        self.tile_location_input.setCurrentIndex(0)
        self.sort_method_input.setCurrentIndex(0)
        self.rxn_number_input.clear()
        self.facs_population_input.clear()
        self.project_input.setCurrentIndex(0)
        self.project_name_input.clear()

        # Clear cDNA metrics
        self.percent_cdna_400bp_input.clear()
        self.cdna_concentration_input.clear()
        self.rna_lib_concentration_input.clear()
        self.atac_lib_concentration_input.clear()

        # Clear library info
        self.cdna_amp_date_input.clear()
        self.atac_prep_date_input.clear()
        self.rna_prep_date_input.clear()
        self.cdna_pcr_cycles_input.clear()
        self.expected_recovery_input.clear()
        self.nuclei_concentration_input.clear()
        self.nuclei_volume_input.clear()

        # Clear indices tab
        self.rna_indices_input.clear()
        self.atac_indices_input.clear()
        self.rna_sizes_input.clear()
        self.atac_sizes_input.clear()
        self.library_cycles_rna_input.clear()
        self.library_cycles_atac_input.clear()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    gui = DataLogGUI()
    gui.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()

## to build on windows: pyinstaller --onefile --windowed --icon=icon.ico --add-data "requirements.txt;." --add-data "sample_name_counter.json;." datalog_gui.py