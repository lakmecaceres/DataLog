import os
import json
import sys
import pyperclip
import dateutil.parser
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Environment Setup ---
if getattr(sys, 'frozen', False):
    script_dir = os.path.dirname(sys.executable)
else:
    script_dir = os.path.dirname(os.path.abspath(__file__))

COUNTER_FILE = os.path.join(script_dir, 'sample_name_counter.json')
workbook_path = os.path.join(script_dir, 'datalog.xlsx')

# Load or initialize counter data
if os.path.exists(COUNTER_FILE):
    with open(COUNTER_FILE, 'r') as f:
        counter_data = json.load(f)
else:
    counter_data = {
        "next_counter": 90,  # Global P number tracker
        "date_info": {},      # Tracks reactions and batches per date
        "amp_counter": {}     # Tracks cDNA amplification batches
    }

# --- Excel File Setup ---
def initialize_excel():
    if os.path.exists(workbook_path):
        wb = load_workbook(workbook_path)
    else:
        wb = Workbook()
        del wb['Sheet']

    headers = [
        'krienen_lab_identifier', 'seq_portal', 'elab_link', 'experiment_start_date',
        'mit_name', 'donor_name', 'tissue_name', 'tissue_name_old',
        'dissociated_cell_sample_name', 'facs_population_plan', 'cell_prep_type',
        'study', 'enriched_cell_sample_container_name', 'expc_cell_capture',
        'port_well', 'enriched_cell_sample_name', 'enriched_cell_sample_quantity_count',
        'barcoded_cell_sample_name', 'library_method', 'cDNA_amplification_method',
        'cDNA_amplification_date', 'amplified_cdna_name', 'cDNA_pcr_cycles',
        'rna_amplification_pass_fail', 'percent_cdna_longer_than_400bp',
        'cdna_amplified_quantity_ng', 'cDNA_library_input_ng', 'library_creation_date',
        'library_prep_set', 'library_name', 'tapestation_avg_size_bp',
        'library_num_cycles', 'lib_quantification_ng', 'library_prep_pass_fail',
        'r1_index', 'r2_index', 'ATAC_index', 'library_pool_name'
    ]

    if 'hmba' not in wb.sheetnames:
        ws = wb.create_sheet('hmba')
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
    else:
        ws = wb['hmba']
        # Check if headers need to be written
        if ws.max_row == 0 or not any(cell.value == 'krienen_lab_identifier' for cell in ws[1]):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

    return wb, ws

# Initialize workbook and worksheet
workbook, worksheet = initialize_excel()
current_row = worksheet.max_row
if current_row == 1 and not any(cell.value for cell in worksheet[1]):
    current_row = 1  # Empty sheet
else:
    current_row += 1  # Append after last row

# --- Style Definitions ---
black_fill = PatternFill(start_color='000000', fill_type='solid')
bold_font = Font(bold=True)

# --- Date Conversion Function ---
def convert(exp_date):
    clean_date = "".join(c for c in exp_date if c.isdigit())
    if len(clean_date) == 6:
        try:
            datetime.strptime(clean_date, '%y%m%d')
            return clean_date
        except ValueError:
            pass
    try:
        parsed_date = dateutil.parser.parse(exp_date)
        return parsed_date.strftime('%y%m%d')
    except ValueError:
        print('Invalid date format. Please try again.')
        return None

# --- Name Mapping Dictionaries ---
name_to_code = {
    "Croissant": "CJ23.56.002",
    "Nutmeg": "CJ23.56.003",
    "Jellybean": "CJ24.56.001",
    "Rambo": "CJ24.56.004",
    "Morel": "CJ24.56.015"
}

tile_location_map = {
    "BRAINSTEM": "BS",
    "BS": "BS",
    "CORTEX": "CX",
    "CX": "CX",
    "CEREBELLUM": "CB",
    "CB": "CB"
}

# --- User Input Collection ---
print("If multiple reactions are run, separate input values using commas.")

# Date input
while True:
    date_input = input('Input the date of the experiment: ')
    current_date = convert(date_input)
    if current_date:
        break

# Initialize counter variables
date_info = counter_data["date_info"]
next_counter = counter_data["next_counter"]

if current_date not in date_info:
    date_info[current_date] = {
        "total_reactions": 0,
        "batches": []
    }

date_entry = date_info[current_date]
existing_total = date_entry["total_reactions"]

# Marmoset name input
while True:
    mit_name_input = input("Input the name of the marmoset: ").strip().title()
    if mit_name_input in name_to_code:
        mit_name = "cj" + mit_name_input
        donor_name = name_to_code[mit_name_input]
        break
    else:
        print("Invalid name. Please enter one of: Croissant, Nutmeg, Jellybean, Rambo, Morel.")

# Slab number input
while True:
    slab_input = input("Input the slab number: ").strip()
    try:
        slab_int = int(slab_input)
        slab = str(slab_int)
        break
    except ValueError:
        print("Invalid slab number. Please enter a numeric value.")

# Tile number input
while True:
    tile_input = input("Input the tile number: ").strip()
    try:
        tile_int = int(tile_input)
        tile = str(tile_int).zfill(2)
        break
    except ValueError:
        print("Invalid tile number. Please enter a numeric value.")

# Hemisphere processing
while True:
    hemisphere = input(
        "Did the tile come from the left hemisphere (LH), right hemisphere (RH), or both? ").strip().lower()
    if hemisphere in ["left", "lh", "right", "rh", "both"]:
        break
    else:
        print("Invalid input. Please enter left/LH, right/RH, or both.")

hemisphere = hemisphere.upper().replace("LEFT", "LH").replace("RIGHT", "RH")
if hemisphere == "RH":
    slab = str(int(slab) + 40).zfill(2)
elif hemisphere == "BOTH":
    slab = str(int(slab) + 90).zfill(2)
else:
    slab = slab.zfill(2)

# Tile location input
while True:
    tile_location_input = input(
        "Is the tile from the Brainstem (BS), Cortex (CX), and/or Cerebellum (CB)? ").strip().upper()
    tile_locations = []
    for part in tile_location_input.replace(" and ", ",").split(","):
        part = part.strip()
        if part in tile_location_map:
            tile_locations.append(tile_location_map[part])
        elif part in ["BS", "CX", "CB"]:
            tile_locations.append(part)
    if tile_locations:
        tile_location_abbr = "-".join(tile_locations)
        break
    else:
        print("Invalid input. Please enter Brainstem/BS, Cortex/CX, or Cerebellum/CB, separated by commas or 'and'.")

# Sort method handling
while True:
    sort_method = input("Input the sort method (pooled/unsorted/DAPI?): ").strip()
    if sort_method.lower() in ["pooled", "unsorted", "dapi"]:
        break
    print("Invalid sort method. Please enter pooled, unsorted, or DAPI.")

sort_method = sort_method.upper() if sort_method.lower() == "dapi" else sort_method

# Reaction number input
while True:
    rxn_number_input = input("Input the number of reactions you ran: ").strip()
    try:
        rxn_number = int(rxn_number_input)
        if rxn_number > 0:
            break
        else:
            print("Please enter a positive integer.")
    except ValueError:
        print("Invalid input. Please enter a numeric value.")

# Update batch calculations
total_reactions_after = existing_total + rxn_number
batches_before = (existing_total + 7) // 8
batches_after = (total_reactions_after + 7) // 8
new_batches_needed = batches_after - batches_before

new_p_numbers = [next_counter + i for i in range(new_batches_needed)]
next_counter += new_batches_needed

all_batches = date_entry["batches"].copy()
all_batches.extend({"p_number": p, "count": 0} for p in new_p_numbers)

port_wells = []
for x in range(rxn_number):
    global_idx = existing_total + x + 1
    batch_idx = (global_idx - 1) // 8
    p_number = all_batches[batch_idx]["p_number"]
    port_well = (global_idx - 1) % 8 + 1
    port_wells.append((p_number, port_well))

# Update counters
date_entry["total_reactions"] = total_reactions_after
date_entry["batches"] = all_batches
counter_data.update({
    "date_info": date_info,
    "next_counter": next_counter
})

# Initialize common values
seq_portal = "no"
elab_link = pyperclip.paste()
tissue_name = f"{donor_name}.{tile_location_abbr}.{slab}.{tile}"
dissociated_cell_sample_name = f'{current_date}_{tissue_name}.Multiome'
cell_prep_type = "nuclei"

# Sorter initials
while True:
    sorter_initials = input("Enter the sorter's first and last initials: ").strip().upper()
    if sorter_initials:
        break
    else:
        print("Initials cannot be empty.")

# FACS population handling
if sort_method.lower() == "pooled":
    while True:
        proportions = input("Enter the proportions of NeuN+/Dneg/Olig2+ (e.g., 70/20/10): ").strip()
        if "/" in proportions:
            proportions_list = proportions.split("/")
            if len(proportions_list) == 3:
                try:
                    proportions_int = [int(p) for p in proportions_list]
                    if sum(proportions_int) == 100:
                        facs_population = "/".join(map(str, proportions_int))
                        break
                    else:
                        print("Proportions must sum to 100.")
                except ValueError:
                    print("Please enter numbers only.")
            else:
                print("Please enter three values separated by slashes.")
        else:
            print("Invalid format. Use slashes to separate values.")
elif sort_method.lower() == "unsorted":
    facs_population = "no_FACS"
else:
    facs_population = "DAPI"

# Study/project handling
while True:
    is_hmba_subcortex = input("Is the sample for the HMBA Subcortex project? (yes/no): ").strip().lower()
    if is_hmba_subcortex in ["yes", "y"]:
        study = "HMBA_CjAtlas_Subcortex"
        break
    elif is_hmba_subcortex in ["no", "n"]:
        study = input("Enter the project name: ").strip()
        if study:
            break
        else:
            print("Project name cannot be empty.")
    else:
        print("Please answer yes or no.")

sorting_status = "PS" if sort_method.lower() in ["pooled", "dapi"] else "PN"
enriched_cell_sample_container_name = f"MPXM_{current_date}_{sorting_status}_{sorter_initials}"

# Cell handling metrics
while True:
    expected_cell_capture_input = input("What is the expected recovery?: ").strip()
    try:
        expected_cell_capture = int(expected_cell_capture_input)
        break
    except ValueError:
        print("Invalid input. Please enter a numeric value.")

while True:
    concentration_input = input("Enter the concentration of nuclei/cells: ").replace(",", "").strip()
    try:
        concentration = float(concentration_input)
        break
    except ValueError:
        print("Invalid input. Please enter a numeric value.")

while True:
    volume_input = input("Enter the volume used (µL): ").strip()
    try:
        volume = float(volume_input)
        break
    except ValueError:
        print("Invalid input. Please enter a numeric value.")

enriched_cell_sample_quantity_count = round(concentration * volume)

# Library date handling
while True:
    cdna_amplification_date_input = input('Input the cDNA amplification date: ')
    cdna_amplification_date = convert(cdna_amplification_date_input)
    if cdna_amplification_date:
        break

while True:
    atac_library_prep_date_input = input("Enter the ATAC library preparation date: ")
    atac_library_prep_date = convert(atac_library_prep_date_input)
    if atac_library_prep_date:
        break

while True:
    rna_library_prep_date_input = input("Enter the cDNA library preparation date: ")
    rna_library_prep_date = convert(rna_library_prep_date_input)
    if rna_library_prep_date:
        break

# cDNA data collection
rna_amplification_pass_fail = "Pass"

while True:
    cdna_pcr_cycles_list = input("Enter the number of cDNA amp cycles for each reaction: ").split(',')
    if len(cdna_pcr_cycles_list) == rxn_number:
        break
    print(f"Please enter {rxn_number} values.")

while True:
    cdna_input = input("Enter the percent of cDNA > 400bp for each reaction: ")
    percent_cdna_long_400bp_list = cdna_input.split(',')
    try:
        percent_cdna_long_400bp_list = [round(float(x.strip())) for x in percent_cdna_long_400bp_list]
        if len(percent_cdna_long_400bp_list) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} numeric values.")

while True:
    cdna_concentration_list = input("Enter the concentration of amplified cDNA (ng/uL) for each reaction: ").split(',')
    try:
        cdna_concentration_list = [float(x.strip()) for x in cdna_concentration_list]
        if len(cdna_concentration_list) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} numeric values.")

cdna_amplified_quantity_ng_list = [conc * 40 for conc in cdna_concentration_list]

# Index handling functions
def convert_index(index):
    index = index.strip().upper()
    if len(index) == 3:
        if index[0].isdigit() and index[1].isdigit() and index[2].isalpha():
            return f"{index[2]}{index[0]}{index[1]}"
        elif index[0].isalpha() and index[1].isdigit() and index[2].isdigit():
            return index
    elif len(index) == 2:
        if index[0].isdigit() and index[1].isalpha():
            return f"{index[1]}0{index[0]}"
        elif index[0].isalpha() and index[1].isdigit():
            return f"{index[0]}0{index[1]}"
    return None

def pad_index(index):
    if len(index) == 2 and index[0].isalpha() and index[1].isdigit():
        return f"{index[0]}0{index[1]}"
    return index

# ATAC indices
while True:
    atac_indices_input = input("Enter the ATAC library indices: ").strip().upper()
    atac_indices = [convert_index(index) for index in atac_indices_input.split(",")]
    if all(atac_indices) and len(atac_indices) == rxn_number:
        atac_indices = [pad_index(index) for index in atac_indices]
        break
    print(f"Please enter {rxn_number} valid ATAC indices (e.g., A1, 2B, C3).")

# RNA indices
while True:
    rna_indices_input = input("Enter the cDNA library indices: ").strip().upper()
    rna_indices = [convert_index(index) for index in rna_indices_input.split(",")]
    if all(rna_indices) and len(rna_indices) == rxn_number:
        rna_indices = [pad_index(index) for index in rna_indices]
        break
    print(f"Please enter {rxn_number} valid cDNA indices (e.g., D4, 5E, F6).")

# Tapestation sizes
while True:
    rna_sizes_input = input(f"Enter the Tapestation average size (bp) for cDNA libraries: ").strip()
    rna_sizes = rna_sizes_input.split(',')
    try:
        rna_sizes = [int(size.strip()) for size in rna_sizes]
        if len(rna_sizes) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} integer values separated by commas.")

while True:
    atac_sizes_input = input(f"Enter the Tapestation average size (bp) for ATAC libraries: ").strip()
    atac_sizes = atac_sizes_input.split(',')
    try:
        atac_sizes = [int(size.strip()) for size in atac_sizes]
        if len(atac_sizes) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} integer values separated by commas.")

# Library cycles
while True:
    library_num_cycles_rna_input = input(f"Enter the number of SI PCR cycles used for cDNA libraries: ").strip()
    try:
        library_num_cycles_rna = [int(x.strip()) for x in library_num_cycles_rna_input.split(',')]
        if len(library_num_cycles_rna) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} integer values separated by commas.")

while True:
    library_num_cycles_atac_input = input(f"Enter the number of SI PCR cycles used for ATAC libraries: ").strip()
    try:
        library_num_cycles_atac = [int(x.strip()) for x in library_num_cycles_atac_input.split(',')]
        if len(library_num_cycles_atac) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} integer values separated by commas.")

# Library quantification
while True:
    lib_quant_rna_input = input(f"Enter the cDNA library concentrations (ng/uL): ").strip()
    try:
        lib_quant_rna = [round(float(x.strip())) for x in lib_quant_rna_input.split(',')]
        if len(lib_quant_rna) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} numeric values separated by commas.")

while True:
    lib_quant_atac_input = input(f"Enter the ATAC library concentrations (ng/uL): ").strip()
    try:
        lib_quant_atac = [round(float(x.strip())) for x in lib_quant_atac_input.split(',')]
        if len(lib_quant_atac) == rxn_number:
            break
    except ValueError:
        pass
    print(f"Please enter {rxn_number} numeric values separated by commas.")

# --- Excel Writing ---
dup_index_counter = {}
headers = [cell.value for cell in worksheet[1]]

for x in range(rxn_number):
    p_number, port_well = port_wells[x]
    barcoded_cell_sample_name = f'P{str(p_number).zfill(4)}_{port_well}'

    for modality in ["RNA", "ATAC"]:
        krienen_lab_identifier = f'{current_date}_HMBA_{mit_name}_Slab{int(slab)}_Tile{int(tile)}_{sort_method}_{modality}{x + 1}'
        enriched_cell_sample_name = f'MPXM_{current_date}_{sorting_status}_{sorter_initials}_{port_well}'
        library_prep_date = rna_library_prep_date if modality == "RNA" else atac_library_prep_date

        if modality == "RNA":
            library_method = "10xMultiome-RSeq"
            library_type = "LPLCXR"
            library_index = rna_indices[x]
        else:
            library_method = "10xMultiome-ASeq"
            library_type = "LPLCXA"
            library_index = atac_indices[x]

        # Update library prep set counter
        key = (library_type, library_prep_date, library_index)
        dup_index_counter[key] = dup_index_counter.get(key, 0) + 1

        library_prep_set = f"{library_type}_{library_prep_date}_{dup_index_counter[key]}"
        library_name = f"{library_prep_set}_{library_index}"

        # Prepare data row
        row_data = [
            krienen_lab_identifier,  # Column 1
            seq_portal,
            elab_link,
            current_date,
            mit_name,
            donor_name,
            tissue_name,
            None,  # tissue_name_old (will be filled black)
            dissociated_cell_sample_name,
            facs_population,
            cell_prep_type,
            study,
            enriched_cell_sample_container_name,
            expected_cell_capture,
            port_well,
            enriched_cell_sample_name,
            enriched_cell_sample_quantity_count,
            barcoded_cell_sample_name,
            library_method,
            "10xMultiome-RSeq" if modality == "RNA" else None,
            cdna_amplification_date if modality == "RNA" else None,
            None,  # amplified_cdna_name (filled conditionally)
            cdna_pcr_cycles_list[x] if modality == "RNA" else None,
            rna_amplification_pass_fail if modality == "RNA" else None,
            percent_cdna_long_400bp_list[x] if modality == "RNA" else None,
            cdna_amplified_quantity_ng_list[x] if modality == "RNA" else None,
            (cdna_amplified_quantity_ng_list[x] * 0.25) if modality == "RNA" else None,
            library_prep_date,
            library_prep_set,
            library_name,
            rna_sizes[x] if modality == "RNA" else atac_sizes[x],
            library_num_cycles_rna[x] if modality == "RNA" else library_num_cycles_atac[x],
            (lib_quant_rna[x] * 35) if modality == "RNA" else (lib_quant_atac[x] * 20),
            "Pass",
            f"SI-TT-{rna_indices[x]}_i7" if modality == "RNA" else None,
            f"SI-TT-{rna_indices[x]}_b(i5)" if modality == "RNA" else None,
            f"SI-NA-{atac_indices[x]}" if modality == "ATAC" else None,
            None  # library_pool_name
        ]

        # Handle amplified_cdna_name for RNA
        if modality == "RNA":
            if current_date not in counter_data["amp_counter"]:
                counter_data["amp_counter"][current_date] = 0
            reaction_count = counter_data["amp_counter"][current_date]
            letter = chr(65 + (reaction_count % 8))
            batch_num_for_amp = (reaction_count // 8) + 1
            row_data[20] = f"APLCXR_{cdna_amplification_date}_{batch_num_for_amp}_{letter}"
            counter_data["amp_counter"][current_date] += 1

        # Write to Excel
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=current_row, column=col_num, value=value)
            # Apply black fill for ATAC empty cells
            if modality == "ATAC" and value is None:
                cell.fill = black_fill

        # Apply black fill to tissue_name_old
        tissue_old_col = headers.index('tissue_name_old') + 1
        worksheet.cell(row=current_row, column=tissue_old_col).fill = black_fill

        current_row += 1

# Adjust column widths
for column in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            cell_value = str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column_letter].width = adjusted_width

# Save outputs
workbook.save(workbook_path)

# --- Persist Data ---
with open(COUNTER_FILE, 'w') as f:
    json.dump(counter_data, f)

print(f"Data successfully appended to {workbook_path}")